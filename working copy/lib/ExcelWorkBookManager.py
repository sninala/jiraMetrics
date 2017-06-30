import ast
import numpy
import string
import datetime
import collections
from Constants import Constants
from dateutil.parser import parse
from openpyxl import Workbook, load_workbook
from ChartManager import ChartManager
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


class ExcelWorkBookManager(object):

    def __init__(self, config):
        self.config = config
        self.min_max_values_for_metric = collections.OrderedDict()
        self.script_executed_for_current_week = False
        self.closed_elapsed_stats_current_week = None
        self.closed_elapsed_grouping_per_project = None
        self.projects_to_calculate_closed_elapsed = None

    def create_empty_workbook(self, output_file_name):
        print "Creating empty workbook as {}".format(output_file_name)
        workbook = Workbook()
        ws = workbook.active
        ws.title = Constants.ROLLUP_SHEET_TITLE
        ws.sheet_properties.tabColor = Constants.ROLLUP_SHEET_COLOR
        # setting Headers in excel Sheets ###
        for header in Constants.ROLLUP_SHEET_HEADERS:
            header_properties = Constants.ROLLUP_SHEET_HEADER_PROPERTIES[header]
            cell_range = header_properties[Constants.CELL_RANGE]
            start_cell, end_cell = cell_range.split(":")
            ws.merge_cells(cell_range)
            ws[start_cell].value = header
            if Constants.SUB_HEADER in header_properties:
                sub_header = header_properties[Constants.SUB_HEADER]
                sub_header_start_column = start_cell[:-1]
                sub_header_end_column = end_cell[:-1]
                length = len(sub_header)
                for i, j in zip(range(length), range(ord(sub_header_start_column), ord(sub_header_end_column) + 1)):
                    ws[chr(j) + Constants.ROLLUP_HEADER_ROWS] = sub_header[i]
        ws.freeze_panes = Constants.ROLLUP_FREEZE_PANE_CELL
        ert_projects = self.get_project_codes()

        for project in ert_projects:
            project_sheet_header = Constants.PROJECT_SHEET_PROPERTIES['SHEET_HEADER']
            project_worksheet = workbook.create_sheet(project)
            project_worksheet.append(project_sheet_header)
            project_worksheet.freeze_panes = Constants.PROJECT_SHEET_PROPERTIES['SHEET_FREEZE_PANE_CELL']
            project_worksheet.sheet_properties.tabColor = self.config.get('PROJECT_COLOR', project)

        closed_elapsed_rollup = workbook.create_sheet(Constants.CLOSED_ELAPSED_ROLLUP_SHEET_TITLE, 0)
        closed_elapsed_rollup_header = Constants.CLOSED_ELAPSED_ROLLUP_SHEET_HEADERS
        closed_elapsed_rollup.append(closed_elapsed_rollup_header)
        closed_elapsed_rollup.freeze_panes = Constants.CLOSED_ELAPSED_ROLLUP_FREEZE_PANE_CELL
        closed_elapsed_rollup.sheet_properties.tabColor = \
            Constants.METRICS[Constants.CLOSED_ELAPSED]['pivot_sheet_color']
        print "Workbook created"
        print "Applying Styles to workbook"
        header_font = Font(name='Calibri', size=12, bold=True)
        side = Side(border_style='thin', color="FF000000")
        color_fill = PatternFill("solid", fgColor="87CEEB")
        wrap_alignment = Alignment(wrap_text=True, vertical="top", horizontal='center')
        for sheet in workbook.get_sheet_names():
            ws = workbook[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom
                    )
                    border.left = border.right = border.top = border.bottom = side
                    cell.alignment = wrap_alignment
                    cell.border = border
                    cell.font = header_font
                    cell.fill = color_fill
        print "Styles applied"
        workbook.save(filename=output_file_name)

    def get_project_names(self):
        project_codes = self.get_project_codes()
        latest_project_names_map = self.get_project_code_mapping_details_for_latest_workbook()
        project_names = []
        for project in project_codes:
            project_name = latest_project_names_map[project]
            project_names.append(project_name)
        return project_names

    @staticmethod
    def is_date(date_string):
        try:
            parse(date_string)
            return True
        except ValueError:
            return False

    def get_project_code_mapping_details_for_old_workbook(self):
        old_workbook_project_name_mapper = collections.OrderedDict()
        old_workbook_project_code_vs_names = self.config.get(
            'BUG_TRACKER', 'old_workbook_project_code_vs_name_map').split(',')
        for item in old_workbook_project_code_vs_names:
            (project_code, project_name) = item.split('=>')
            old_workbook_project_name_mapper[project_code.strip()] = project_name.strip()
        return old_workbook_project_name_mapper

    def get_project_code_mapping_details_for_latest_workbook(self):
        latest_workbook_project_name_mapper = collections.OrderedDict()
        latest_project_code_vs_names = self.config.get('BUG_TRACKER', 'project_code_vs_name_map').split(',')
        for item in latest_project_code_vs_names:
            (project_code, project_name) = item.split('=>')
            latest_workbook_project_name_mapper[project_code.strip()] = project_name.strip()
        return latest_workbook_project_name_mapper

    def extract_data_from_old_file_and_insert_into_new_file(self, old_workbook_file_name, out_put_file_name):
        old_workbook_project_name_mapper = self.get_project_code_mapping_details_for_old_workbook()
        latest_workbook_project_name_mapper = self.get_project_code_mapping_details_for_latest_workbook()
        old_workbook_project_names = old_workbook_project_name_mapper.values()
        print "Loading the old workbook {}".format(old_workbook_file_name)
        old_workbook = load_workbook(old_workbook_file_name, data_only=True)
        old_workbook_rollup = old_workbook[Constants.ROLLUP_SHEET_TITLE]
        old_workbook_project_sheet_data = collections.OrderedDict()
        old_workbook_rollup_sheet_data = collections.OrderedDict()
        print "Extracting Rollup data from {} workbook and loading into {} workbook".format(
            old_workbook_file_name, out_put_file_name
        )
        for row in old_workbook_rollup.iter_rows():
            if row[0].row == 1:
                continue
            old_workbook_project_sheet_data[row[2].value + '##' + str(row[3].value)] = "##".join(
                [str(row[4].value), str(row[7].value), str(row[10].value)])
            data = []
            for i in range(4, 19):
                data.append(row[i].value)
            data = [str(cell_data) for cell_data in data]
            old_workbook_rollup_sheet_data[row[2].value + '##' + str(row[3].value)] = "##".join(data)
        dates = set()
        keys = old_workbook_project_sheet_data.keys()
        for key in keys:
            project, date = key.split("##")
            if ExcelWorkBookManager.is_date(date):
                dates.add(date)
        dates = sorted(dates, key=lambda x: datetime.datetime.strptime(x, '%Y-%m-%d %H:%M:%S'))
        latest_workbook = load_workbook(out_put_file_name)
        # populate the project specific sheets
        for project in old_workbook_project_names:
            project1 = project
            if project1 == 'Expert':
                project1 = 'EXPRT'
            elif project1 == 'ePRO':
                project1 = 'EPR'
            ws = latest_workbook[project1.upper()]
            for date in dates:
                key = project + '##' + date
                if key in old_workbook_project_sheet_data.keys():
                    (New, InProgress, Closed) = old_workbook_project_sheet_data[key].split('##')
                    project_sheet_max_row = ws.max_row
                    date1 = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
                    week_of_year = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S').strftime("%W-%Y")
                    if project_sheet_max_row == 1:
                        diff1 = diff2 = diff3 = 0
                    else:
                        diff1 = "=C{0}-C{1}".format(project_sheet_max_row + 1, project_sheet_max_row)
                        diff2 = "=E{0}-E{1}".format(project_sheet_max_row + 1, project_sheet_max_row)
                        diff3 = "=G{0}-G{1}".format(project_sheet_max_row + 1, project_sheet_max_row)
                    row = [week_of_year, date1, int(New), diff1, int(InProgress), diff2, int(Closed), diff3]
                    ws.append(row)
        # populate the Rollup
        latest_workbook_rollup = latest_workbook['Rollup']
        for date in dates:
            for project_code, project in old_workbook_project_name_mapper.iteritems():
                key = project + '##' + date
                project1 = latest_workbook_project_name_mapper[project_code]
                if key in old_workbook_rollup_sheet_data.keys():
                    date1 = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
                    values = old_workbook_rollup_sheet_data[key].split('##')
                    int_values = []
                    for value in values:
                        try:
                            value = int(value)
                        except ValueError:
                            value = str(value)
                        int_values.append(value)
                    row = [project1, date1] + int_values
                    latest_workbook_rollup.append(row)
        print "Extracting ClosedElapsed data from {} workbook and loading into {} workbook".format(
            old_workbook_file_name, out_put_file_name
        )
        old_workbook_closed_elapsed_rollup = old_workbook['Pivot0-ClosedElapsed']
        latest_workbook_closed_elapsed_rollup = latest_workbook[Constants.CLOSED_ELAPSED_ROLLUP_SHEET_TITLE]
        old_workbook_closed_elapsed_sheet_data = collections.OrderedDict()

        for row in old_workbook_closed_elapsed_rollup.iter_rows():
            if row[0].row == 1:
                continue
            project = row[0].value
            run_date = row[1].value
            avg_days_elapsed = row[2].value
            max_days_elapsed = row[3].value
            min_days_elapsed = row[4].value
            median_days_elapsed = row[5].value
            if project and run_date and avg_days_elapsed and max_days_elapsed \
                    and min_days_elapsed and median_days_elapsed:
                old_workbook_closed_elapsed_sheet_data[project + '##' + str(run_date)] = "##".join(
                    [str(avg_days_elapsed), str(max_days_elapsed), str(min_days_elapsed), str(median_days_elapsed)])
        dates = set()
        keys = old_workbook_closed_elapsed_sheet_data.keys()
        for key in keys:
            project, date = key.split("##")
            if ExcelWorkBookManager.is_date(date):
                dates.add(date)
        dates = sorted(dates, key=lambda x: datetime.datetime.strptime(x, '%Y-%m-%d %H:%M:%S'))
        project_names = self.get_project_names()
        for date in dates:
            for project in project_names:
                key = project + '##' + date
                if key in old_workbook_closed_elapsed_sheet_data.keys():
                    date1 = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
                    values = old_workbook_closed_elapsed_sheet_data[key].split('##')
                    numbers = []
                    for value in values:
                        try:
                            value = ast.literal_eval(value)
                        except ValueError:
                            value = str(value)
                        numbers.append(value)
                    row = [project, date1] + numbers
                    latest_workbook_closed_elapsed_rollup.append(row)
        latest_workbook.save(filename=out_put_file_name)
        print "Data Extraction Completed"

    def is_workbook_already_has_data_for_current_week(self, workbook, run_date_str):
        ert_projects = self.get_project_codes()
        project_worksheet = workbook[ert_projects[0]]
        if project_worksheet.max_row == 1:
            script_executed_for_current_week = False
        else:
            last_run_date = project_worksheet.cell(
                row=project_worksheet.max_row, column=project_worksheet.min_column + 1
            ).value
            last_run_date = last_run_date.strftime("%m/%d/%Y")
            script_executed_for_current_week = (last_run_date == run_date_str)
        return script_executed_for_current_week

    def populate_latest_metrics_from_jira_for_date(self, program_run_date, jira_api, out_put_file_name):
        run_date_str = program_run_date.strftime("%m/%d/%Y")
        run_date_yyyy_mm_dd = program_run_date.strftime("%Y-%m-%d")
        run_week = program_run_date.strftime("%W-%Y")
        latest_workbook_project_name_mapper = self.get_project_code_mapping_details_for_latest_workbook()
        print "Getting the latest metrics from jira for {}".format(run_date_str)
        workbook = load_workbook(out_put_file_name)
        sheets = workbook.get_sheet_names()
        # populating data for project Sheets #
        rollup_sheet = workbook['Rollup']
        rollup_sheet_max_row = rollup_sheet.max_row
        self.script_executed_for_current_week = self.is_workbook_already_has_data_for_current_week(
            workbook, run_date_str)
        ert_projects = self.get_project_codes()
        if self.script_executed_for_current_week:
            rollup_sheet_max_row = (rollup_sheet_max_row - len(ert_projects))

        rollup_sheet_rows = []
        rollup_index = 1
        last_week_result = dict()
        current_week_results = dict()
        for project in ert_projects:
            print "working on Metrics for {} project ".format(project)
            is_project_newly_added = False
            if project in sheets:
                worksheet = workbook[project]
            else:
                project_sheet_header = Constants.PROJECT_SHEET_PROPERTIES['SHEET_HEADER']
                worksheet = workbook.create_sheet(project)
                worksheet.append(project_sheet_header)
                worksheet.freeze_panes = Constants.PROJECT_SHEET_PROPERTIES['SHEET_FREEZE_PANE_CELL']
                worksheet.sheet_properties.tabColor = self.config.get('PROJECT_COLOR', project)
                is_project_newly_added = True
            row = []
            jql_queries = collections.OrderedDict(self.config.items('JQL'))
            if self.config.has_section(project + '_JQL'):
                jql_queries = collections.OrderedDict(self.config.items(project + '_JQL'))
            project_sheet_max_row = worksheet.max_row
            if self.script_executed_for_current_week:
                project_sheet_max_row = (project_sheet_max_row - 1)
            for status, query in jql_queries.iteritems():
                if status.lower() == 'closedelapsed':
                    continue
                status = string.capwords(status)
                query = query.replace('__PROJECTNAME__', project)
                query = query.replace('__CURRENTDATE__', run_date_yyyy_mm_dd)
                response = jira_api.get_response_from_jira(query)
                count = response['total']
                # time.sleep(10)
                row.append(count)
                print status, count
                current_week_results[project + '-' + status] = count
                (last_run_value, diff) = (0, 0)
                if project_sheet_max_row == 1:
                    pass
                elif status == 'New':
                    last_run_value = worksheet['C' + str(project_sheet_max_row)].value
                    diff = "=C{0}-C{1}".format(project_sheet_max_row + 1, project_sheet_max_row)
                elif status == 'In Progress':
                    last_run_value = worksheet['E' + str(project_sheet_max_row)].value
                    diff = "=E{0}-E{1}".format(project_sheet_max_row + 1, project_sheet_max_row)
                elif status == 'Closed':
                    last_run_value = worksheet['G' + str(project_sheet_max_row)].value
                    diff = "=G{0}-G{1}".format(project_sheet_max_row + 1, project_sheet_max_row)
                row.append(diff)
                last_week_result[project + '-' + status] = last_run_value
            row = [run_week, program_run_date] + row
            if self.script_executed_for_current_week:
                index = 0
                for col in range(worksheet.min_column, worksheet.max_column + 1):
                    worksheet.cell(row=project_sheet_max_row + 1, column=col, value=row[index])
                    index += 1
            else:
                worksheet.append(row)
            if rollup_sheet_max_row == 2 or is_project_newly_added:
                rollup_sheet_rows.append(
                    [latest_workbook_project_name_mapper[project], program_run_date,
                     current_week_results[project + '-New'], last_week_result[project + '-New'], 0,
                     current_week_results[project + '-In Progress'], last_week_result[project + '-In Progress'], 0,
                     current_week_results[project + '-Closed'], last_week_result[project + '-Closed'], 0,
                     "=C{0}+F{0}".format(rollup_sheet_max_row + rollup_index), 0, 0,
                     "=I{0}+L{0}".format(rollup_sheet_max_row + rollup_index), 0, 0]
                )
            else:
                rollup_sheet_rows.append(
                    [latest_workbook_project_name_mapper[project], program_run_date,
                     current_week_results[project + '-New'], last_week_result[project + '-New'],
                     "=C{0}-D{0}".format(rollup_sheet_max_row + rollup_index),
                     current_week_results[project + '-In Progress'], last_week_result[project + '-In Progress'],
                     "=F{0}-G{0}".format(rollup_sheet_max_row + rollup_index),
                     current_week_results[project + '-Closed'], last_week_result[project + '-Closed'],
                     "=I{0}-J{0}".format(rollup_sheet_max_row + rollup_index),
                     "=C{0}+F{0}".format(rollup_sheet_max_row + rollup_index),
                     "=D{0}+G{0}".format(rollup_sheet_max_row + rollup_index),
                     "=L{0}-M{0}".format(rollup_sheet_max_row + rollup_index),
                     "=I{0}+L{0}".format(rollup_sheet_max_row + rollup_index),
                     "=J{0}+M{0}".format(rollup_sheet_max_row + rollup_index),
                     "=O{0}-P{0}".format(rollup_sheet_max_row + rollup_index)]
                )
            rollup_index += 1
        # populate data for Rollup Sheet ###
        if self.script_executed_for_current_week:
            index = 1
            for row in rollup_sheet_rows:
                for col in range(rollup_sheet.min_column, rollup_sheet.max_column + 1):
                    rollup_sheet.cell(
                        row=rollup_sheet_max_row + index, column=col, value=rollup_sheet_rows[index - 1][col - 1]
                    )
                index += 1
        else:
            for row in rollup_sheet_rows:
                rollup_sheet.append(row)
        print "Getting latest Closed Elapsed Metrics From Jira"
        closed_elapsed_stats = collections.OrderedDict()
        closed_elapsed_grouping_per_project = collections.OrderedDict()
        closed_elapsed_rollup_sheet = workbook[Constants.CLOSED_ELAPSED_ROLLUP_SHEET_TITLE]
        closed_elapsed_rollup_sheet_max_row = closed_elapsed_rollup_sheet.max_row
        if self.script_executed_for_current_week:
            closed_elapsed_rollup_sheet_max_row = (closed_elapsed_rollup_sheet_max_row - len(ert_projects))
        closed_elapsed_rollup_rows = []
        projects_to_calculate_closed_elapsed = list()
        closed_elapsed_field_name = Constants.CLOSED_ELAPSED_FIELD_IN_JIRA_API_RESPONSE
        s = set()
        for project in ert_projects:
            total_items = 1
            response_count = 0
            current_project_closed_elapsed_stats = []
            closed_elapsed_grouping_per_project[project] = {}
            query = self.config.get('JQL', 'ClosedElapsed')
            query = query.replace('__PROJECTNAME__', project)
            query = query.replace('__CURRENTDATE__', run_date_yyyy_mm_dd)
            while response_count < total_items:
                response = jira_api.get_response_from_jira(query, str(response_count), closed_elapsed_field_name)
                total_items = response['total']
                response_count = response_count + len(response['issues'])
                print "Extracted {} records out of {} for project {} from Jira".format(
                    response_count, total_items, project)
                issues = response['issues']
                for issue in issues:
                    closed_elapsed_value = int(issue['fields'][closed_elapsed_field_name])
                    if closed_elapsed_value and (project not in s):
                        s.add(project)
                        projects_to_calculate_closed_elapsed.append(project)
                    current_project_closed_elapsed_stats.append(closed_elapsed_value)
                    # grouping tickets by closed_elapsed value
                    if closed_elapsed_value in closed_elapsed_grouping_per_project[project]:
                        closed_elapsed_grouping_per_project[project][closed_elapsed_value] = \
                            closed_elapsed_grouping_per_project[project][closed_elapsed_value] + 1
                    else:
                        closed_elapsed_grouping_per_project[project][closed_elapsed_value] = 1

            current_project_metrics = self.get_closed_elapsed_metrics(current_project_closed_elapsed_stats)
            row = [latest_workbook_project_name_mapper[project], program_run_date] + current_project_metrics
            closed_elapsed_stats[project] = row
            if not row[2] and not row[3] and not row[4] and not row[5]:
                pass
            else:
                closed_elapsed_rollup_rows.append(row)
        if self.script_executed_for_current_week:
            index = 1
            for row in closed_elapsed_rollup_rows:
                for col in range(closed_elapsed_rollup_sheet.min_column, closed_elapsed_rollup_sheet.max_column + 1):
                    closed_elapsed_rollup_sheet.cell(
                        row=closed_elapsed_rollup_sheet_max_row + index,
                        column=col, value=closed_elapsed_rollup_rows[index - 1][col - 1]
                    )
                index += 1
        else:
            for row in closed_elapsed_rollup_rows:
                closed_elapsed_rollup_sheet.append(row)
        self.set_closed_elapsed_metrics_for_current_week(closed_elapsed_stats, closed_elapsed_grouping_per_project)
        self.set_projects_to_calculate_closed_elapsed(projects_to_calculate_closed_elapsed)
        self.apply_styles_to_the_workbook(workbook)
        workbook.save(filename=out_put_file_name)

    def set_closed_elapsed_metrics_for_current_week(
            self, closed_elapsed_stats_current_week, closed_elapsed_grouping_per_project):
        self.closed_elapsed_stats_current_week = closed_elapsed_stats_current_week
        self.closed_elapsed_grouping_per_project = closed_elapsed_grouping_per_project

    def get_closed_elapsed_metrics_for_current_week(self):
        return [self.closed_elapsed_stats_current_week, self.closed_elapsed_grouping_per_project]

    def set_projects_to_calculate_closed_elapsed(self, projects):
        self.projects_to_calculate_closed_elapsed = projects

    def get_projects_to_calculate_closed_elapsed(self):
        return self.projects_to_calculate_closed_elapsed

    def get_pivot_metrics_from_work_book(self, workbook, weekly_metric_type):
        ert_projects = self.get_project_codes()
        weekly_total_metrics = collections.OrderedDict()
        weekly_totals_closed = collections.OrderedDict()
        inprogress_weekly_totals = collections.OrderedDict()
        new_weekly_totals = collections.OrderedDict()
        result = None
        for project in ert_projects:
            ws = workbook[project]
            for row in ws.iter_rows():
                rows = []
                for cell in row:
                    if cell.row == 1:
                        continue
                    rows.append(cell.value)
                if rows:
                    run_date = rows[1]
                    run_date = run_date.strftime("%m/%d/%Y")
                    weekly_total_metrics[project + '#' + run_date] = '#'.join(str(v) for v in rows[2:])
                    weekly_totals_closed[project + '#' + run_date + '#' + Constants.STATUS_CLOSED] = rows[6]
                    inprogress_weekly_totals[project + '#' + run_date + '#' + Constants.STATUS_INPROGRESS] = rows[4]
                    new_weekly_totals[project + '#' + run_date + '#' + Constants.STATUS_NEW] = rows[2]

        if weekly_metric_type == Constants.ALL_TICKETS_WEEKLY_TOTALS:
            result = weekly_total_metrics
        elif weekly_metric_type == Constants.CLOSED_WEEKLY_TOTALS:
            result = weekly_totals_closed
        elif weekly_metric_type == Constants.IN_PROGRESS_WEEKLY_TOTALS:
            result = inprogress_weekly_totals
        elif weekly_metric_type == Constants.NEW_WEEKLY_TOTALS:
            result = new_weekly_totals
        return result

    def apply_styles_to_the_workbook(self, workbook):
        sheets = [Constants.ROLLUP_SHEET_TITLE] + [Constants.CLOSED_ELAPSED_ROLLUP_SHEET_TITLE]
        sheets = sheets + self.get_project_codes()
        side = Side(border_style='thin', color="FF000000")
        for sheet in sheets:
            ws = workbook[sheet]
            min_row = 2
            if sheet == Constants.ROLLUP_SHEET_TITLE:
                min_row = 3
            for row in range(min_row, ws.max_row + 1):
                for column in range(ws.min_column, ws.max_column + 1):
                    cell = ws.cell(row=row, column=column)
                    if column == 2:
                        cell.number_format = 'M/D/YYYY'
                    if sheet == Constants.CLOSED_ELAPSED_ROLLUP_SHEET_TITLE and column in (3, 4, 5, 6):
                        cell.number_format = '0'
                    border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom
                    )
                    border.left = border.right = border.top = border.bottom = side
                    cell.border = border

    def create_or_update_pivot_table_for(self, metric_name, out_put_file_name, program_run_date):
        workbook = load_workbook(out_put_file_name)
        sheets = workbook.get_sheet_names()
        pivot_sheet_name = Constants.METRICS[metric_name]['pivot_sheet_name']
        pivot_sheet_position = Constants.METRICS[metric_name]['pivot_sheet_position']
        pivot_sheet_color = Constants.METRICS[metric_name]['pivot_sheet_color']
        if not (pivot_sheet_name in sheets):
            print "Creating Sheet {}".format(pivot_sheet_name)
            pivots_worksheet = workbook.create_sheet(pivot_sheet_name, pivot_sheet_position)
            pivots_worksheet.sheet_properties.tabColor = pivot_sheet_color
        else:
            pivots_worksheet = workbook.get_sheet_by_name(pivot_sheet_name)
            self.clear_the_content_in_a_sheet(workbook, pivots_worksheet, out_put_file_name)
        self.create_pivot_tables_for(metric_name, pivots_worksheet, workbook, program_run_date)
        workbook.save(filename=out_put_file_name)

    @staticmethod
    def clear_the_content_in_a_sheet(workbook, sheet, out_put_file_name):
        for row in sheet.iter_rows(row_offset=1):
            for cell in row:
                cell.value = None
        workbook.save(filename=out_put_file_name)

    def create_pivot_tables_for(self, metric_name, pivots_worksheet, workbook, program_run_date):
        print "Creating Pivot tables for metric {}".format(metric_name)
        if metric_name == Constants.ALL_TICKETS_WEEKLY_TOTALS:
            self.create_all_tickets_weekly_total_pivot_tables(workbook, pivots_worksheet)
        elif metric_name == Constants.CLOSED_WEEKLY_TOTALS:
            self.create_closed_weekly_totals_pivot_tables(workbook, pivots_worksheet)
        elif metric_name == Constants.CLOSED_WEEKLY_CHANGE:
            self.create_closed_weekly_change_pivot_tables(workbook, pivots_worksheet)
        elif metric_name == Constants.IN_PROGRESS_WEEKLY_TOTALS:
            self.create_inprogress_weekly_total_pivot_tables(workbook, pivots_worksheet)
        elif metric_name == Constants.IN_PROGRESS_WEEKLY_CHANGE:
            self.create_inprogress_weekly_change_pivot_tables(workbook, pivots_worksheet)
        elif metric_name == Constants.NEW_WEEKLY_TOTALS:
            self.create_new_weekly_total_pivot_tables(workbook, pivots_worksheet)
        elif metric_name == Constants.NEW_WEEKLY_CHANGE:
            self.create_new_weekly_change_pivot_tables(workbook, pivots_worksheet)
        elif metric_name == Constants.CLOSED_ELAPSED:
            self.create_closed_elapsed_pivot_tables(workbook, pivots_worksheet, program_run_date)

    def create_all_tickets_weekly_total_pivot_tables(self, workbook, pivots_worksheet):
        weekly_total_metrics = self.get_pivot_metrics_from_work_book(workbook, Constants.ALL_TICKETS_WEEKLY_TOTALS)
        run_dates = ExcelWorkBookManager.get_run_dates_from_metrics(weekly_total_metrics)
        project_rows = []
        for run_date in run_dates:
            total = 0
            for project in self.get_project_codes():
                key = project + '#' + run_date
                if key in weekly_total_metrics:
                    (new, diff1, in_progress, diff2, closed, diff3) = weekly_total_metrics[key].split('#')
                else:
                    (new, in_progress, closed) = (0, 0, 0)
                project_total = int(new) + int(in_progress) + int(closed)
                total = total + project_total
            project_rows.append((run_date, total))
        change_in_growth = []
        for i in range(1, len(project_rows)):
            change_in_growth.append((project_rows[i][0], project_rows[i][1] - project_rows[i - 1][1]))

        project_rows = self.restrict_pivot_data_based_on_rollup_weeks(project_rows)
        change_in_growth = self.restrict_pivot_data_based_on_rollup_weeks(change_in_growth)

        change_in_growth = [('Date', 'Weekly Growth in Tickets')] + change_in_growth
        col = 1
        row = 1
        for data_row in change_in_growth:
            for index, value in enumerate(data_row):
                pivots_worksheet.cell(row=row, column=col+index).value = data_row[index]
            row = row + 1
        col = 5
        row = 1
        project_rows = [('Date', 'Sum of All Tickets')] + project_rows
        for data_row in project_rows:
            for index, value in enumerate(data_row):
                pivots_worksheet.cell(row=row, column=col+index).value = data_row[index]
            row = row + 1

    def restrict_pivot_data_based_on_rollup_weeks(self, data_array):
        number_of_rollup_weeks = self.config.get('OUTPUT', 'number_of_rollup_weeks')
        if number_of_rollup_weeks and number_of_rollup_weeks.lower() != 'all':
            number_of_rollup_weeks = int(number_of_rollup_weeks)
            if number_of_rollup_weeks and (len(data_array) > number_of_rollup_weeks):
                offset = len(data_array) - number_of_rollup_weeks
                data_array = data_array[offset:]
        return data_array

    def populate_project_specific_metrics_from_closed_elapsed_rollup(self, workbook, pivots_worksheet):
        elapsed_rollup_sheet = workbook[Constants.CLOSED_ELAPSED_ROLLUP_SHEET_TITLE]
        closed_elapsed_stats = collections.OrderedDict()
        for row in elapsed_rollup_sheet.iter_rows():
            if row[0].row == 1:
                continue
            project = row[0].value
            run_date = row[1].value
            avg_days_elapsed = row[2].value
            max_days_elapsed = row[3].value
            min_days_elapsed = row[4].value
            median_days_elapsed = row[5].value
            if project and run_date and avg_days_elapsed and max_days_elapsed \
                    and min_days_elapsed and median_days_elapsed:
                closed_elapsed_stats[project + '##' + str(run_date)] = "##".join(
                    [str(avg_days_elapsed), str(max_days_elapsed), str(min_days_elapsed), str(median_days_elapsed)])
        dates = set()
        keys = closed_elapsed_stats.keys()
        for key in keys:
            project, date = key.split("##")
            if self.is_date(date):
                dates.add(date)
        dates = sorted(dates, key=lambda x: datetime.datetime.strptime(x, '%Y-%m-%d %H:%M:%S'))
        dates = self.restrict_pivot_data_based_on_rollup_weeks(dates)
        project_names = self.get_project_names()
        col = 6
        for project in project_names:
            project_rows = list()
            for date in dates:
                key = project + '##' + date
                project, run_date = key.split('##')
                if key in closed_elapsed_stats.keys():
                    date1 = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
                    values = closed_elapsed_stats[key].split('##')
                    project_row = [date1, float(values[0]), float(values[3]), int(values[1])]
                    project_rows.append(project_row)
            if project_rows:
                row = 1
                pivots_worksheet.cell(row=row, column=col, value="Project:")
                pivots_worksheet.cell(row=row, column=col + 1, value=project)
                row = row + 1
                pivots_worksheet.cell(row=row, column=col, value="Run Date")
                pivots_worksheet.cell(row=row, column=col + 1, value="Average")
                pivots_worksheet.cell(row=row, column=col + 2, value="Median")
                pivots_worksheet.cell(row=row, column=col + 3, value="Max")
                row = row + 1
                for data_row in project_rows:
                    pivots_worksheet.cell(row=row, column=col, value=data_row[0])
                    pivots_worksheet.cell(row=row, column=col).number_format = 'M/D/YYYY'
                    pivots_worksheet.cell(row=row, column=col + 1, value=data_row[1])
                    pivots_worksheet.cell(row=row, column=col + 1).number_format = '0'
                    pivots_worksheet.cell(row=row, column=col+2, value=data_row[2])
                    pivots_worksheet.cell(row=row, column=col + 2).number_format = '0'
                    pivots_worksheet.cell(row=row, column=col+3, value=data_row[3])
                    row = row + 1
                col = col + 5

    def populate_closed_elapsed_grouping_per_project(
            self, workbook, pivots_worksheet, closed_elapsed_grouping_per_project):
        days_elapsed = list()
        ert_projects = self.get_projects_to_calculate_closed_elapsed()
        project_name_map = self.get_project_code_mapping_details_for_latest_workbook()
        project_names = list()
        col = 6
        for project in ert_projects:
            col = col + 5
            project_names.append(project_name_map[project])
            days_elapsed_per_project = closed_elapsed_grouping_per_project[project].keys()
            days_elapsed.extend(days_elapsed_per_project)
        days_elapsed = set(days_elapsed)
        days_elapsed = sorted(days_elapsed)
        row = 1
        pivots_worksheet.cell(row=row, column=col, value="Number of Jira Tickets Per Elapsed Day")
        row = row + 1
        pivots_worksheet.cell(row=row, column=col, value="Days Elapsed")
        for i in range(len(project_names)):
            pivots_worksheet.cell(row=row, column=col + i + 1, value=project_names[i])
        pivots_worksheet.cell(row=row, column=col + len(project_names)+1, value="Total")
        row = row + 1

        for days_elapsed_value in days_elapsed:
            pivots_worksheet.cell(row=row, column=col, value=days_elapsed_value)
            col_index = 1
            total = 0
            for project in ert_projects:
                grouping_for_project = None
                if days_elapsed_value in closed_elapsed_grouping_per_project[project]:
                    grouping_for_project = closed_elapsed_grouping_per_project[project][days_elapsed_value]
                    total = total + grouping_for_project
                pivots_worksheet.cell(row=row, column=col + col_index, value=grouping_for_project)
                col_index = col_index + 1
            pivots_worksheet.cell(row=row, column=col + col_index, value=total)
            row = row + 1

    def create_closed_elapsed_pivot_tables(self, workbook, pivots_worksheet, program_run_date):
        closed_elapsed_stats_current_week = self.get_closed_elapsed_metrics_for_current_week()[0]
        closed_elapsed_grouping_per_project = self.get_closed_elapsed_metrics_for_current_week()[1]
        run_date_str = program_run_date.strftime("%m/%d/%Y")
        ert_projects = self.get_projects_to_calculate_closed_elapsed()
        row = 1
        col = 1
        pivots_worksheet.cell(row=row, column=col, value="Run Date")
        pivots_worksheet.cell(row=row, column=col+1, value=run_date_str)
        row = row + 1
        pivots_worksheet.cell(row=row, column=col, value="Project")
        pivots_worksheet.cell(row=row, column=col + 1, value="Average")
        pivots_worksheet.cell(row=row, column=col + 2, value="Max")
        pivots_worksheet.cell(row=row, column=col + 3, value="Median")
        row = row + 1
        for project in ert_projects:
            metrics_for_project = closed_elapsed_stats_current_week[project]
            pivot_row = [metrics_for_project[0], metrics_for_project[2], metrics_for_project[3], metrics_for_project[5]]
            if not pivot_row[1] and not pivot_row[2] and not pivot_row[3]:
                pass
            else:
                pivots_worksheet.cell(row=row, column=col, value=pivot_row[0])
                pivots_worksheet.cell(row=row, column=col + 1, value=pivot_row[1])
                pivots_worksheet.cell(row=row, column=col + 1).number_format = '0'
                pivots_worksheet.cell(row=row, column=col + 2, value=pivot_row[2])
                pivots_worksheet.cell(row=row, column=col + 3, value=pivot_row[3])
                pivots_worksheet.cell(row=row, column=col + 3).number_format = '0'
                row = row + 1
        self.populate_project_specific_metrics_from_closed_elapsed_rollup(workbook, pivots_worksheet)
        self.populate_closed_elapsed_grouping_per_project(
            workbook, pivots_worksheet, closed_elapsed_grouping_per_project)

    def create_weekly_total_pivot_table_for(self, metric_name, workbook, pivots_worksheet):
        status = None
        if metric_name == Constants.IN_PROGRESS_WEEKLY_TOTALS:
            status = Constants.STATUS_INPROGRESS
        elif metric_name == Constants.NEW_WEEKLY_TOTALS:
            status = Constants.STATUS_NEW
        elif metric_name == Constants.CLOSED_WEEKLY_TOTALS:
            status = Constants.STATUS_CLOSED
        weekly_totals = self.get_pivot_metrics_from_work_book(workbook, metric_name)
        run_dates = ExcelWorkBookManager.get_run_dates_from_metrics(weekly_totals)
        ert_projects = self.get_project_codes()
        ert_project_names = self.get_project_names()
        header_row = ['Run Date'] + ert_project_names
        weekly_total_pivot_rows = list()
        for date in run_dates:
            row = [date]
            for project in ert_projects:
                key = project + '#' + date + '#' + status
                if key in weekly_totals:
                    weekly_totals_for_project = weekly_totals[key]
                else:
                    weekly_totals_for_project = 0
                row.append(weekly_totals_for_project)
            weekly_total_pivot_rows.append(row)

        weekly_total_pivot_rows = self.restrict_pivot_data_based_on_rollup_weeks(weekly_total_pivot_rows)
        weekly_total_pivot_rows.insert(0, header_row)
        col = 1
        row = 1
        for data_row in weekly_total_pivot_rows:
            for index, value in enumerate(data_row):
                pivots_worksheet.cell(row=row, column=col + index).value = data_row[index]
            row = row + 1
        self.calculate_project_min_max_values_for_metric(metric_name, weekly_total_pivot_rows)

    def calculate_project_min_max_values_for_metric(self, metric_name, weekly_total_pivot_rows):
        row = 1
        projects = list()
        all_values = list()
        self.min_max_values_for_metric[metric_name] = collections.OrderedDict()
        max_value = 0
        max_value_project_name = None
        for data_row in weekly_total_pivot_rows:
            for index, value in enumerate(data_row):
                if row == 1 and index >= 1:
                    project = data_row[index]
                    projects.append(project)
                    self.min_max_values_for_metric[metric_name][project] = list()
                elif row > 1 and index >= 1:
                    value = data_row[index]
                    if value > max_value:
                        max_value = value
                        max_value_project_name = projects[index-1]
                    self.min_max_values_for_metric[metric_name][projects[index-1]].append(value)
                    all_values.append(value)
            row = row + 1
        # print max_value_project_name

        if metric_name in [Constants.IN_PROGRESS_WEEKLY_TOTALS, Constants.NEW_WEEKLY_TOTALS,
                           Constants.CLOSED_WEEKLY_TOTALS]:
            max_project_values = self.min_max_values_for_metric[metric_name][max_value_project_name]
            max_project_values = sorted(set(max_project_values))
            max_project_max_value = max_project_values[len(max_project_values) - 1]
            max_project_min_value = max_project_values[0]
            max_rounded_value = int(50 * round(max_project_max_value / 50)) + 100
            min_rounded_value = int(50 * round(max_project_min_value / 50)) - 50
        else:
            all_values = sorted(set(all_values))
            min_all_projects = all_values[0]
            max_all_projects = all_values[len(all_values) - 1]
            max_rounded_value = int(5 * round(max_all_projects / 5)) + 10
            min_rounded_value = int(5 * round(min_all_projects / 5)) - 5
        diff = max_rounded_value - min_rounded_value

        for project in projects:
            if project == max_value_project_name:
                self.min_max_values_for_metric[metric_name][project] = [min_rounded_value, max_rounded_value]
            else:
                if metric_name in [Constants.IN_PROGRESS_WEEKLY_TOTALS, Constants.NEW_WEEKLY_TOTALS,
                                   Constants.CLOSED_WEEKLY_TOTALS]:
                    values = self.min_max_values_for_metric[metric_name][project]
                    values = sorted(set(values))
                    project_max_value = values[len(values) - 1]
                    project_max_value = int(50 * round(project_max_value / 50)) + 100
                    project_min_value = project_max_value - diff
                    if project_min_value < 0:
                        project_max_value = diff
                        project_min_value = 0
                    else:
                        project_min_value = int(50 * round(project_min_value / 50))
                    self.min_max_values_for_metric[metric_name][project] = [project_min_value, project_max_value]
                else:
                    values = self.min_max_values_for_metric[metric_name][project]
                    values = sorted(set(values))
                    project_max_value = values[len(values) - 1]
                    project_max_value = max_rounded_value
                    project_min_value = min_rounded_value
                    self.min_max_values_for_metric[metric_name][project] = [project_min_value, project_max_value]

    def create_weekly_change_pivot_table_for(self, metric_name, workbook, pivots_worksheet):
        metric_name_total = None
        status = None
        if metric_name == Constants.IN_PROGRESS_WEEKLY_CHANGE:
            metric_name_total = Constants.IN_PROGRESS_WEEKLY_TOTALS
            status = Constants.STATUS_INPROGRESS
        elif metric_name == Constants.NEW_WEEKLY_CHANGE:
            metric_name_total = Constants.NEW_WEEKLY_TOTALS
            status = Constants.STATUS_NEW
        elif metric_name == Constants.CLOSED_WEEKLY_CHANGE:
            metric_name_total = Constants.CLOSED_WEEKLY_TOTALS
            status = Constants.STATUS_CLOSED
        weekly_totals = self.get_pivot_metrics_from_work_book(workbook, metric_name_total)
        run_dates = ExcelWorkBookManager.get_run_dates_from_metrics(weekly_totals)
        ert_projects = self.get_project_codes()
        ert_project_names = self.get_project_names()
        header_row = ['Run Date'] + ert_project_names
        weekly_change_pivot_rows = list()
        for i in range(1, len(run_dates)):
            weekly_change_row = [run_dates[i]]
            for project in ert_projects:
                key1 = project + '#' + run_dates[i] + '#' + status
                key2 = project + '#' + run_dates[i - 1] + '#' + status
                if key1 in weekly_totals:
                    current_value = weekly_totals[key1]
                else:
                    current_value = 0
                if key2 in weekly_totals:
                    old_value = weekly_totals[key2]
                else:
                    old_value = 0
                weekly_change_row.append(current_value - old_value)
            weekly_change_pivot_rows.append(weekly_change_row)
        weekly_change_pivot_rows = self.restrict_pivot_data_based_on_rollup_weeks(weekly_change_pivot_rows)
        weekly_change_pivot_rows.insert(0, header_row)
        col = 1
        row = 1
        for data_row in weekly_change_pivot_rows:
            for index, value in enumerate(data_row):
                pivots_worksheet.cell(row=row, column=col + index).value = data_row[index]
            row = row + 1
        self.calculate_project_min_max_values_for_metric(metric_name, weekly_change_pivot_rows)

    def create_inprogress_weekly_total_pivot_tables(self, workbook, pivots_worksheet):
        self.create_weekly_total_pivot_table_for(Constants.IN_PROGRESS_WEEKLY_TOTALS, workbook, pivots_worksheet)

    def create_new_weekly_total_pivot_tables(self, workbook, pivots_worksheet):
        self.create_weekly_total_pivot_table_for(Constants.NEW_WEEKLY_TOTALS, workbook, pivots_worksheet)

    def create_closed_weekly_totals_pivot_tables(self, workbook, pivots_worksheet):
        self.create_weekly_total_pivot_table_for(Constants.CLOSED_WEEKLY_TOTALS, workbook, pivots_worksheet)

    def create_closed_weekly_change_pivot_tables(self, workbook, pivots_worksheet):
        self.create_weekly_change_pivot_table_for(Constants.CLOSED_WEEKLY_CHANGE, workbook, pivots_worksheet)

    def create_inprogress_weekly_change_pivot_tables(self, workbook, pivots_worksheet):
        self.create_weekly_change_pivot_table_for(Constants.IN_PROGRESS_WEEKLY_CHANGE, workbook, pivots_worksheet)

    def create_new_weekly_change_pivot_tables(self, workbook, pivots_worksheet):
        self.create_weekly_change_pivot_table_for(Constants.NEW_WEEKLY_CHANGE, workbook, pivots_worksheet)

    @staticmethod
    def get_run_dates_from_metrics(metrics):
        keys = metrics.keys()
        run_dates_set = set()
        run_dates = list()
        for key in keys:
            values = key.split('#')
            project, run_date = values[0], values[1]
            if run_date not in run_dates_set:
                run_dates_set.add(run_date)
                run_dates.append(run_date)
        return run_dates

    @staticmethod
    def get_maximum_row(worksheet, column_number):
        max_row = 0
        for row_num in range(1, worksheet.max_row + 2):
            if not worksheet.cell(row=row_num, column=column_number).value:
                max_row = row_num - 1
                break
        return max_row

    def get_project_codes(self):
        project_codes = [project.strip() for project in self.config.get('BUG_TRACKER', 'projects').split(",")]
        return project_codes

    @staticmethod
    def get_closed_elapsed_metrics(array_list):
        median = numpy.median(numpy.array(array_list))
        avg = numpy.average(array_list)
        # avg = int(round(avg))
        array_list1 = sorted(array_list)
        max_days_elapsed = array_list1[len(array_list1) - 1]
        min_days_elapsed = array_list1[0]
        return [avg, max_days_elapsed, min_days_elapsed, median]

    def update_charts_for(self, metrics, out_put_file_name):
        workbook = load_workbook(out_put_file_name)
        sheets = workbook.get_sheet_names()
        for metric_name in metrics:
            pivots_worksheet = workbook.get_sheet_by_name(metrics[metric_name]['pivot_sheet_name'])
            charts_sheet_name = metrics[metric_name]['charts_sheet_name']
            charts_sheet_position = metrics[metric_name]['charts_sheet_position']
            charts_sheet_color = metrics[metric_name]['charts_sheet_color']
            if not (charts_sheet_name in sheets):
                print "Creating Sheet {}".format(charts_sheet_name)
                charts_worksheet = workbook.create_sheet(charts_sheet_name, charts_sheet_position)
                charts_worksheet.sheet_properties.tabColor = charts_sheet_color
            else:
                charts_worksheet = workbook.get_sheet_by_name(charts_sheet_name)
            chart_manager = ChartManager(self.config, pivots_worksheet, charts_worksheet)
            self.draw_charts_for(metric_name, chart_manager)

        workbook.save(filename=out_put_file_name)

    def draw_charts_for(self, metric_name, chart_manager):
        if metric_name == Constants.ALL_TICKETS_WEEKLY_TOTALS:
            weekly_total_date_column_number = 5
            weekly_total_value_column_number = 6
            weekly_total_max_row = self.get_maximum_row(chart_manager.data_sheet, 5)
            barchart_properties = dict()
            barchart_properties['title'] = Constants.METRICS[metric_name]['chart_weekly_total_title']
            barchart_properties['logarithmic_y_axis'] = False
            barchart_properties['data_min_column'] = weekly_total_value_column_number
            barchart_properties['data_min_row'] = 1
            barchart_properties['data_max_column'] = weekly_total_value_column_number
            barchart_properties['data_max_row'] = weekly_total_max_row
            barchart_properties['cats_min_column'] = weekly_total_date_column_number
            barchart_properties['cats_min_row'] = 2
            barchart_properties['cats_max_column'] = weekly_total_value_column_number
            barchart_properties['cats_max_row'] = weekly_total_max_row
            barchart_properties['trendline'] = True
            barchart_properties['data_labels'] = True
            barchart_properties['cell'] = 'A2'
            barchart_properties['projects'] = []
            chart_manager.draw_barchart(barchart_properties)
            growth_change_date_column_number = 1
            growth_change_value_column_number = 2
            weekly_growth_max_row = self.get_maximum_row(chart_manager.data_sheet, 1)
            linechart_properties = dict()
            linechart_properties['logarithmic_y_axis'] = False
            linechart_properties['title'] = Constants.METRICS[metric_name]['chart_weekly_growth_title']
            linechart_properties['data_min_column'] = growth_change_value_column_number
            linechart_properties['data_min_row'] = 1
            linechart_properties['data_max_column'] = growth_change_value_column_number
            linechart_properties['data_max_row'] = weekly_growth_max_row
            linechart_properties['cats_min_column'] = growth_change_date_column_number
            linechart_properties['cats_min_row'] = 2
            linechart_properties['cats_max_column'] = growth_change_date_column_number
            linechart_properties['cats_max_row'] = weekly_growth_max_row
            linechart_properties['trendline'] = True
            linechart_properties['data_labels'] = False
            linechart_properties['cell'] = 'A30'
            linechart_properties['projects'] = []
            linechart_properties['statistics'] = []
            chart_manager.draw_linechart(linechart_properties)
        elif (metric_name == Constants.CLOSED_WEEKLY_TOTALS)\
                or (metric_name == Constants.IN_PROGRESS_WEEKLY_TOTALS)\
                or (metric_name == Constants.NEW_WEEKLY_TOTALS):
            title = Constants.METRICS[metric_name]['chart_weekly_total_title']
            barchart_properties = dict()
            barchart_properties['title'] = title + ' - All'
            barchart_properties['logarithmic_y_axis'] = False
            barchart_properties['data_min_column'] = 2
            barchart_properties['data_min_row'] = 1
            barchart_properties['data_max_column'] = chart_manager.data_sheet.max_column
            barchart_properties['data_max_row'] = chart_manager.data_sheet.max_row
            barchart_properties['cats_min_column'] = 1
            barchart_properties['cats_min_row'] = 2
            barchart_properties['cats_max_column'] = chart_manager.data_sheet.max_column
            barchart_properties['cats_max_row'] = chart_manager.data_sheet.max_row
            barchart_properties['trendline'] = False
            barchart_properties['data_labels'] = False
            barchart_properties['cell'] = 'A2'
            barchart_properties['projects'] = self.get_project_codes()
            chart_manager.draw_barchart(barchart_properties)
            self.draw_charts_for_metrics_at_project_level(chart_manager, title, metric_name, "barchart")
        elif (metric_name == Constants.CLOSED_WEEKLY_CHANGE)\
                or (metric_name == Constants.IN_PROGRESS_WEEKLY_CHANGE)\
                or (metric_name == Constants.NEW_WEEKLY_CHANGE):
            title = Constants.METRICS[metric_name]['chart_weekly_change_title']
            linechart_properties = dict()
            linechart_properties['title'] = title
            linechart_properties['logarithmic_y_axis'] = False
            linechart_properties['data_min_column'] = 2
            linechart_properties['data_min_row'] = 1
            linechart_properties['data_max_column'] = chart_manager.data_sheet.max_column
            linechart_properties['data_max_row'] = chart_manager.data_sheet.max_row
            linechart_properties['cats_min_column'] = 1
            linechart_properties['cats_min_row'] = 2
            linechart_properties['cats_max_column'] = chart_manager.data_sheet.max_column
            linechart_properties['cats_max_row'] = chart_manager.data_sheet.max_row
            linechart_properties['trendline'] = False
            linechart_properties['data_labels'] = False
            linechart_properties['cell'] = 'A2'
            linechart_properties['projects'] = self.get_project_codes()
            linechart_properties['statistics'] = []
            chart_manager.draw_linechart(linechart_properties)
            self.draw_charts_for_metrics_at_project_level(chart_manager, title, metric_name, "linechart")
        elif metric_name == Constants.CLOSED_ELAPSED:
            max_row = self.get_maximum_row(chart_manager.data_sheet, 1)
            title = Constants.METRICS[metric_name]['chart_current_week_title']
            linechart_properties = dict()
            linechart_properties['title'] = title
            linechart_properties['logarithmic_y_axis'] = True
            linechart_properties['data_min_column'] = 2
            linechart_properties['data_min_row'] = 2
            linechart_properties['data_max_column'] = 4
            linechart_properties['data_max_row'] = max_row
            linechart_properties['cats_min_column'] = 1
            linechart_properties['cats_min_row'] = 3
            linechart_properties['cats_max_column'] = 4
            linechart_properties['cats_max_row'] = max_row
            linechart_properties['trendline'] = False
            linechart_properties['data_labels'] = True
            linechart_properties['cell'] = 'A2'
            linechart_properties['projects'] = []
            linechart_properties['statistics'] = Constants.CLOSED_ELAPSED_STATISTICS
            chart_manager.draw_linechart(linechart_properties)
            self.draw_charts_for_closed_elapsed_metric_per_project(chart_manager, "linechart")
            self.draw_charts_for_closed_elapsed_metric_per_elapsed_day(chart_manager, "barchart")

    def draw_charts_for_metrics_at_project_level(self, chart_manager, title, metric_name, chart_type):
        data_sheet = chart_manager.data_sheet
        ert_projects = self.get_project_codes()
        project_name_mapper = self.get_project_code_mapping_details_for_latest_workbook()
        cell_index = 30
        for index, project in enumerate(ert_projects):
            project_name = project_name_mapper[project]
            chart_properties = dict()
            chart_properties['logarithmic_y_axis'] = False
            chart_properties['title'] = title + " - " + project_name
            chart_properties['data_min_column'] = index + 2
            chart_properties['data_min_row'] = 1
            chart_properties['data_max_column'] = index + 2
            chart_properties['data_max_row'] = data_sheet.max_row
            chart_properties['cats_min_column'] = 1
            chart_properties['cats_min_row'] = 2
            chart_properties['cats_max_column'] = chart_manager.data_sheet.max_column
            chart_properties['cats_max_row'] = chart_manager.data_sheet.max_row
            chart_properties['trendline'] = True
            chart_properties['data_labels'] = True
            chart_properties['projects'] = [project]
            chart_properties['statistics'] = []
            chart_properties['cell'] = 'A' + str(cell_index)
            chart_properties['y_axis_min_value'] = self.min_max_values_for_metric[metric_name][project_name][0]
            chart_properties['y_axis_max_value'] = self.min_max_values_for_metric[metric_name][project_name][1]
            cell_index += 30
            if chart_type == "linechart":
                chart_manager.draw_linechart(chart_properties)
            elif chart_type == "barchart":
                chart_manager.draw_barchart(chart_properties)

    def draw_charts_for_closed_elapsed_metric_per_project(self, chart_manager, chart_type):
        project_name_mapper = self.get_project_code_mapping_details_for_latest_workbook()
        data_sheet = chart_manager.data_sheet
        ert_projects = self.get_projects_to_calculate_closed_elapsed()
        cell_index = 30
        # starting at column 6 because, first 5 columns are used by closed elapsed metrics for all projects
        col = 6
        for project in ert_projects:
            max_row = self.get_maximum_row(data_sheet, col)
            chart_properties = dict()
            chart_properties['logarithmic_y_axis'] = True
            chart_properties['title'] = "Closed Elapsed for " + project_name_mapper[project]
            chart_properties['data_min_column'] = col + 1
            chart_properties['data_min_row'] = 2
            chart_properties['data_max_column'] = col + 3
            chart_properties['data_max_row'] = max_row
            chart_properties['cats_min_column'] = col
            chart_properties['cats_min_row'] = 3
            chart_properties['cats_max_column'] = col + 3
            chart_properties['cats_max_row'] = max_row
            chart_properties['trendline'] = False
            chart_properties['data_labels'] = True
            chart_properties['projects'] = []
            chart_properties['statistics'] = ["Average", "Median", "Max"]
            chart_properties['cell'] = 'A' + str(cell_index)
            cell_index += 30
            if chart_type == "linechart":
                chart_manager.draw_linechart(chart_properties)
            col = col + 5

    def draw_charts_for_closed_elapsed_metric_per_elapsed_day(self, chart_manager, chart_type):
        data_sheet = chart_manager.data_sheet
        ert_projects = self.get_projects_to_calculate_closed_elapsed()
        # 6 columns for closed elapsed for all projects
        col = 6 + (5 * len(ert_projects))
        # first 30 rows to display the chart for closed Elapsed metrics for all projects
        # 30 rows for each project closed elapsed metric
        cell_index = 30 + (30 * len(ert_projects))
        max_row = self.get_maximum_row(data_sheet, col)
        chart_properties = dict()
        chart_properties['logarithmic_y_axis'] = True
        chart_properties['title'] = "The Number of Jira Tickets per Elapsed Day"
        chart_properties['data_min_column'] = col + len(ert_projects) + 1
        chart_properties['data_min_row'] = 2
        chart_properties['data_max_column'] = col + len(ert_projects) + 1
        chart_properties['data_max_row'] = max_row
        chart_properties['cats_min_column'] = col
        chart_properties['cats_min_row'] = 3
        chart_properties['cats_max_column'] = col + len(ert_projects) + 1
        chart_properties['cats_max_row'] = max_row
        chart_properties['trendline'] = False
        chart_properties['data_labels'] = False
        chart_properties['stacked'] = True

        chart_properties['projects'] = []
        chart_properties['cell'] = 'A' + str(cell_index)
        if chart_type == "barchart":
            chart_manager.draw_barchart(chart_properties)
