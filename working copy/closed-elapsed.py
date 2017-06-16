import requests
import urllib
import datetime
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook, load_workbook
import os, re, numpy
import collections
from dateutil.parser import parse
import datetime

def get_response_from_jira(query, startAt):
    base_url = "https://jira.ert.com/jira/rest/api/2/search?jql="
    username = "sheth.veeradasari"
    password = "eRT1234_"
    query_string = urllib.quote_plus(query)
    response = requests.get(base_url + query_string + '&maxResults=-1&startAt='+startAt + '&fields=customfield_10950', auth=HTTPBasicAuth(username, password))
    if response.status_code == 200:
        response_json = response.json()
        return response_json
    else:
        raise Exception("Unable get response from Jira - {}".format(response.reason))

def is_date(date_string):
    try:
        parse(date_string)
        return True
    except ValueError:
        return False

def get_closed_elapsed_metrics(arrayList):
    median = numpy.median(numpy.array(arrayList))
    avg = numpy.average(arrayList)
    arrayList1 = sorted(arrayList)
    max_days_elapsed = arrayList1[len(arrayList1)-1]
    min_days_elapsed = arrayList1[0]
    return [avg, max_days_elapsed, min_days_elapsed, median]


    
    


if __name__ == "__main__":
  ert_projects = ['EXPRT', 'EPR', 'MPORT', 'RCVS', 'SPOR']
  ert_projects = ['EXPRT']
  CURRENT_DIRECTORY = os.path.dirname(os.path.realpath(__file__))
  program_run_date = datetime.date.today() - datetime.timedelta(days=0)
  run_date_yyyy_mm_dd = program_run_date.strftime("%Y-%m-%d")
  '''
  old_workbook_file_name = None
  for filename in os.listdir(CURRENT_DIRECTORY):
      match = re.match('(From.*?.xlsm)', filename, re.I)
      if match:
          old_workbook_file_name = match.group(0)
  if old_workbook_file_name:
    old_workbook_file_name = os.path.join(CURRENT_DIRECTORY, old_workbook_file_name)
    if os.path.exists(old_workbook_file_name):
        old_workbook = load_workbook(old_workbook_file_name, data_only=True)
        print "Extracting Closed-Elapsed data from {} workbook ".format(old_workbook_file_name)
        old_workbook_rollup = old_workbook['Pivot0-ClosedElapsed']
        old_workbook_closed_elapsed_sheet_data = collections.OrderedDict()
        for row in old_workbook_rollup.iter_rows():
            if row[0].row == 1:
                continue
            project = row[0].value
            run_date = row[1].value
            avg_days_elapsed = row[2].value
            max_days_elapsed = row[3].value
            min_days_elapsed = row[4].value
            median_days_elapsed = row[5].value
            if project and run_date and avg_days_elapsed and max_days_elapsed and min_days_elapsed and median_days_elapsed:
                old_workbook_closed_elapsed_sheet_data[project + '##' + str(run_date)] = "##".join(
                    [str(avg_days_elapsed), str(max_days_elapsed), str(min_days_elapsed), str(median_days_elapsed)])
        dates = set()
        keys = old_workbook_closed_elapsed_sheet_data.keys()
        for key in keys:
            project, date = key.split("##")
            if is_date(date):
                dates.add(date)
        dates = sorted(dates, key=lambda x: datetime.datetime.strptime(x, '%Y-%m-%d %H:%M:%S'))
        

  '''


  wb = Workbook()
  ws = wb.active
  closed_elapsed_stats = collections.OrderedDict()
  ws.append(["project", "Average", "Max", "Median"])
  for project in ert_projects:
    total_items = 1
    respone_count = 0
    closed_elapsed_stats[project] = []
    time = datetime.datetime.now()
    while(respone_count < total_items):
        query = "project in (__PROJECTNAME__) AND status in (Resolved, Closed) AND createdDate > 2015-01-01 AND createdDate<= __CURRENTDATE__"
        query = query.replace('__PROJECTNAME__', project)
        query = query.replace('__CURRENTDATE__', run_date_yyyy_mm_dd)
        response = get_response_from_jira(query, str(respone_count))
        total_items = response['total']
        respone_count = respone_count + len(response['issues'])
        print project, respone_count, total_items
        issues = response['issues']
        for issue in issues:
            print issue
            print issue["key"], issue['fields']['customfield_10950']
            if issue['fields']['customfield_10950']:
                closed_elapsed_stats[project].append(issue['fields']['customfield_10950'])
            else:
                closed_elapsed_stats[project].append(issue['fields']['customfield_10950'])
    #current_project_metrics = get_closed_elapsed_metrics(closed_elapsed_stats[project])
    #row = [project, current_project_metrics[0],  current_project_metrics[1], current_project_metrics[3]]
    #ws.append(row)
    time1=datetime.datetime.now()
    print time1 - time
  wb.save(filename=r"D:\closed-elapsed.xlsx")

