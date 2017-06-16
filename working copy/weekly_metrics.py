from openpyxl import load_workbook
import collections
import datetime
from openpyxl.utils import coordinate_from_string
from openpyxl.chart import BarChart, LineChart, Series, Reference
from openpyxl.chart.chartspace import DataTable


excelFileName = r"D:\Jira_Metrics_latest.xlsx"
ertProjects = ['EXPRT', 'EPR', 'MPORT', 'RCVS', 'SPOR', 'CRQST']
currentDate = datetime.date.today() - datetime.timedelta(days=0)
currentWeek = currentDate.strftime("%W-%Y")
currentDate_YYYY_MM_DD = currentDate.strftime("%Y-%m-%d")
currentDate = currentDate.strftime("%m/%d/%Y")


def get_pivot_metrics(weekly_metric_type):
    weekly_total_metrics = collections.OrderedDict()
    weekly_totals_closed = collections.OrderedDict()
    for project in ertProjects:
        ws = workBook[project]
        for row in ws.iter_rows():
            rowList = []
            for cell in row:
                if cell.row == 1:
                    continue
                rowList.append(cell.value)
            if rowList:
                weekly_total_metrics[project + '#' + rowList[1]] = '#'.join(str(v) for v in rowList[2:])
                weekly_totals_closed[project + '#' + rowList[1] + '#closed'] = rowList[6]
    if weekly_metric_type == "weekly_totals":
        result = weekly_total_metrics
    elif weekly_metric_type == "weekly_closed_totals":
        result = weekly_totals_closed
    return result


def create_pivot_tables(workBook, pivots_worksheet):
    weekly_total_metrics = get_pivot_metrics("weekly_totals")
    keys = weekly_total_metrics.keys()
    rundatesSet = set()
    rundates = list()
    for key in keys:
        project, rundate = key.split('#')
        if rundate not in rundatesSet:
            rundatesSet.add(rundate)
            rundates.append(rundate)
    #project_rows = [('Date', 'Total')]
    project_rows = []
    for rundate in rundates:
        total = 0
        for project in ertProjects:
            (New, diff1, InProgess, diff2, closed, diff3) = weekly_total_metrics[project + '#' + rundate].split('#')
            projectTotal = int(New) + int(InProgess) + int(closed)
            total = total + projectTotal
        project_rows.append((rundate, total))
    change_in_growth=[]
    for i in range(1,len(project_rows)):
        change_in_growth.append((project_rows[i][0], project_rows[i][1] - project_rows[i-1][1]))
    print change_in_growth

def update_weekly_total_pivot_tables(workBook, pivots_worksheet):
    metrics = collections.OrderedDict()
    for project in ertProjects:
        ws = workBook[project]
        row = ws.max_row
        latest_row = []
        for col in range(ws.min_column, ws.max_column + 1):
            latest_row.append((ws.cell(row=row, column=col).value))
        metrics[project + '#' + latest_row[1]] = '#'.join(str(v) for v in latest_row[2:])
    project, rundate = metrics.keys()[0].split('#')
    total = 0
    for project in ertProjects:
        (New, diff1, InProgess, diff2, closed, diff3) = metrics[project + '#' + rundate].split('#')
        projectTotal = int(New) + int(InProgess) + int(closed)
        total = (total + projectTotal)
    newRow = (rundate, total)
    max_row = 0;
    print pivots_worksheet.max_row
    for row_num in range(1, pivots_worksheet.max_row + 2):
        if not pivots_worksheet.cell(row=row_num, column=1).value:
            max_row = row_num
            break
    print max_row
    for col in pivots_worksheet.iter_cols(min_row=1, min_col=1, max_col=1):
        for cell in col:
            print cell.value, cell.row

def update_closed_weekly_total_pivot_tables(workBook, closed_weekly_pivots_worksheet):
    latest_weekly_totals_closed = collections.OrderedDict()
    for project in ertProjects:
        ws = workBook[project]
        row = ws.max_row
        latest_row = []
        for col in range(ws.min_column, ws.max_column + 1):
            latest_row.append((ws.cell(row=row, column=col).value))
        latest_weekly_totals_closed[project + '#' + latest_row[1] + '#closed'] = latest_row[6]
    project, rundate, status = latest_weekly_totals_closed.keys()[0].split('#')
    total = 0
    closed_counts = []
    for project in ertProjects:
        key = latest_weekly_totals_closed[project + '#' + rundate + '#closed']
        closed_tickets_for_project = latest_weekly_totals_closed[key]
        closed_tickets_for_project = int(closed_tickets_for_project)
        closed_counts.append(closed_tickets_for_project)
    closed_counts = [rundate] + closed_counts
    max_row = closed_weekly_pivots_worksheet.max_row

    for row in range(2, max_row + 1):
        old_rundate = closed_weekly_pivots_worksheet.cell(row=row, column=1).value
        if(old_rundate == currentDate):
            for col in range(1, closed_weekly_pivots_worksheet + 1):
                closed_weekly_pivots_worksheet.cell(row=max_row + 1, column=col, value=closed_counts[col])
            return

    for col in range(1, closed_weekly_pivots_worksheet + 1):
        closed_weekly_pivots_worksheet.cell(row=max_row + 1, column=col, value=closed_counts[col])



def create_closed_weekly_totals_pivot_tables(workBook, closed_weekly_pivots_worksheet):
    weekly_closed_total_metrics = get_pivot_metrics('weekly_closed_totals')
    keys = weekly_closed_total_metrics.keys()
    rundatesSet = set()
    rundates = list()
    for key in keys:
        project, rundate, status = key.split('#')
        if rundate not in rundatesSet:
            rundatesSet.add(rundate)
            rundates.append(rundate)

    header_row = ['run_date']  + ertProjects
    closed_weekly_pivots_worksheet.append(header_row)
    for date in rundates:
        row = [date]
        for project in ertProjects:
            key = project + '#' + date + '#' + 'closed'
            total_closed_for_date_for_project = weekly_closed_total_metrics[key]
            row.append(total_closed_for_date_for_project)
        closed_weekly_pivots_worksheet.append(row)
    workBook.save(filename=excelFileName)




(weekly_total_pivots_sheet, weekly_total_charts_sheet) = ('Pivot-WeeklyTotals', 'Charts-WeeklyTotals')
(closed_weekly_total_pivots_sheet, closed_weekly_total_charts_sheet) = ('Pivot-Weekly-Closed-Totals', 'Charts-Weekly-Closed-Charts')

workBook = load_workbook(excelFileName)
sheets = workBook.get_sheet_names()
if not (closed_weekly_total_pivots_sheet in sheets and closed_weekly_total_charts_sheet in sheets):
    print "Creating Sheet {}".format(closed_weekly_total_pivots_sheet)
    closed_weekly_pivots_worksheet = workBook.create_sheet(closed_weekly_total_pivots_sheet, 2)
    closed_weekly_pivots_worksheet.sheet_properties.tabColor = "1072BA"
    print "Creating Sheet {}".format(closed_weekly_total_charts_sheet)
    closed_weekly_charts_worksheet = workBook.create_sheet(closed_weekly_total_charts_sheet, 2)
    closed_weekly_charts_worksheet.sheet_properties.tabColor = "1072BA"
    create_closed_weekly_totals_pivot_tables(workBook, closed_weekly_pivots_worksheet)
else:
    closed_weekly_pivots_worksheet = workBook.get_sheet_by_name(closed_weekly_total_pivots_sheet)
    closed_weekly_charts_worksheet = workBook.get_sheet_by_name(closed_weekly_total_charts_sheet)
    update_weekly_total_pivot_tables(workBook, closed_weekly_pivots_worksheet)



chart1 = BarChart()
chart1.height = 12
chart1.width = 30
chart1.style = 10
chart1.title = "Weekly Total - All Tickets"
chart1.y_axis.title = 'Total'
chart1.x_axis.title = 'Run Date'
data = Reference(closed_weekly_pivots_worksheet, min_col=2, min_row=1,
                 max_row=closed_weekly_pivots_worksheet.max_row, max_col=closed_weekly_pivots_worksheet.max_column)
cats = Reference(closed_weekly_pivots_worksheet, min_col=1, min_row=2, max_row=closed_weekly_pivots_worksheet.max_row)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
#chart1.series[0].trendline = Trendline()
#chart1.series[0].trendline.trendlineType = 'linear'
#chart1.dataLabels = DataLabelList()
#chart1.dataLabels.showVal = True

closed_weekly_charts_worksheet.add_chart(chart1, "A1")

chart2 = BarChart()
chart2.height = 12
chart2.width = 30
chart2.style = 10
chart2.title = "Weekly Total - All Tickets"
chart2.y_axis.title = 'Total'
chart2.x_axis.title = 'Run Date'
data = Reference(closed_weekly_pivots_worksheet, min_col=3, min_row=1,
                 max_row=closed_weekly_pivots_worksheet.max_row, max_col=3)
cats = Reference(closed_weekly_pivots_worksheet, min_col=1, min_row=2, max_row=closed_weekly_pivots_worksheet.max_row)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
chart2.shape = 4
#chart1.series[0].trendline = Trendline()
#chart1.series[0].trendline.trendlineType = 'linear'
#chart1.dataLabels = DataLabelList()
#chart1.dataLabels.showVal = True

closed_weekly_charts_worksheet.add_chart(chart2, "A20")


workBook.save(filename=excelFileName)

'''
if not (weekly_total_pivots_sheet in sheets and weekly_total_charts_sheet in sheets):
    print "I am in "
    print "Creating Sheet {}".format(weekly_total_pivots_sheet)
    pivots_worksheet = workBook.create_sheet(weekly_total_pivots_sheet, 0)
    pivots_worksheet.sheet_properties.tabColor = "1072BA"
    print "Creating Sheet {}".format(weekly_total_charts_sheet)
    charts_worksheet = workBook.create_sheet(weekly_total_charts_sheet, 0)
    charts_worksheet.sheet_properties.tabColor = "1072BA"
    create_pivot_tables(workBook, pivots_worksheet)
else:
    pivots_worksheet = workBook.get_sheet_by_name(weekly_total_pivots_sheet)
    update_weekly_total_pivot_tables(workBook, pivots_worksheet)
'''

