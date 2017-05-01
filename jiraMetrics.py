"""
Created on Mar 3, 2017
@author: Siva Ninala
"""
import os
import re
import datetime
import collections
import string
import time
import sys
from dateutil.parser import parse
from ConfigParser import SafeConfigParser
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Series, Reference
from openpyxl.chart.trendline import Trendline
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from lib.JiraAPIHandler import JiraAPIHandler


def is_date(date_string):
    try:
        parse(date_string)
        return True
    except ValueError:
        return False

def get_pivot_metrics(weekly_metric_type):
    weekly_total_metrics = collections.OrderedDict()
    weekly_totals_closed = collections.OrderedDict()
    for project in ertProjects:
        ws = workBook[project]
        for row in ws.iter_rows():
            rowList = []
            for cell in row:
                if cell.row == 1:
                    continue
                rowList.append(cell.value)
            if rowList:
                weekly_total_metrics[project + '#' + rowList[1]] = '#'.join(str(v) for v in rowList[2:])
                weekly_totals_closed[project + '#' + rowList[1] + '#closed'] = rowList[6]
    if weekly_metric_type == "weekly_totals":
        result = weekly_total_metrics
    elif weekly_metric_type == "weekly_closed_totals":
        result = weekly_totals_closed
    return result


def update_closed_weekly_total_pivot_tables(workBook, closed_weekly_pivots_worksheet):
    latest_weekly_totals_closed = collections.OrderedDict()
    for project in ertProjects:
        ws = workBook[project]
        row = ws.max_row
        latest_row = []
        for col in range(ws.min_column, ws.max_column + 1):
            latest_row.append((ws.cell(row=row, column=col).value))
        latest_weekly_totals_closed[project + '#' + latest_row[1] + '#closed'] = latest_row[6]
    project, rundate, status = latest_weekly_totals_closed.keys()[0].split('#')
    total = 0
    closed_counts = []
    for project in ertProjects:
        key = latest_weekly_totals_closed[project + '#' + rundate + '#closed']
        closed_tickets_for_project = latest_weekly_totals_closed[key]
        closed_tickets_for_project = int(closed_tickets_for_project)
        closed_counts.append(closed_tickets_for_project)
    closed_counts = [rundate] + closed_counts
    max_row = closed_weekly_pivots_worksheet.max_row

    for row in range(2, max_row + 1):
        old_rundate = closed_weekly_pivots_worksheet.cell(row=row, column=1).value
        if(old_rundate == currentDate):
            for col in range(1, closed_weekly_pivots_worksheet + 1):
                closed_weekly_pivots_worksheet.cell(row=max_row + 1, column=col, value=closed_counts[col])
            return

    for col in range(1, closed_weekly_pivots_worksheet + 1):
        closed_weekly_pivots_worksheet.cell(row=max_row + 1, column=col, value=closed_counts[col])



def create_closed_weekly_totals_pivot_tables(workBook, closed_weekly_pivots_worksheet):
    weekly_closed_total_metrics = get_pivot_metrics('weekly_closed_totals')
    keys = weekly_closed_total_metrics.keys()
    rundatesSet = set()
    rundates = list()
    for key in keys:
        project, rundate, status = key.split('#')
        if rundate not in rundatesSet:
            rundatesSet.add(rundate)
            rundates.append(rundate)

    header_row = ['run_date']  + ertProjects
    closed_weekly_pivots_worksheet.append(header_row)
    for date in rundates:
        row = [date]
        for project in ertProjects:
            key = project + '#' + date + '#' + 'closed'
            total_closed_for_date_for_project = weekly_closed_total_metrics[key]
            row.append(total_closed_for_date_for_project)
        closed_weekly_pivots_worksheet.append(row)



def create_or_update_closed_weekly_total_charts(excelFileName, currentDate):
    (weekly_total_pivots_sheet, weekly_total_charts_sheet) = ('Pivot-WeeklyTotals', 'Charts-WeeklyTotals')
    (closed_weekly_total_pivots_sheet, closed_weekly_total_charts_sheet) = (
    'Pivot-Weekly-Closed-Totals', 'Charts-Weekly-Closed-Charts')

    workBook = load_workbook(excelFileName)
    sheets = workBook.get_sheet_names()
    if not (closed_weekly_total_pivots_sheet in sheets and closed_weekly_total_charts_sheet in sheets):
        print "Creating Sheet {}".format(closed_weekly_total_pivots_sheet)
        closed_weekly_pivots_worksheet = workBook.create_sheet(closed_weekly_total_pivots_sheet, 2)
        closed_weekly_pivots_worksheet.sheet_properties.tabColor = "1072BA"
        print "Creating Sheet {}".format(closed_weekly_total_charts_sheet)
        closed_weekly_charts_worksheet = workBook.create_sheet(closed_weekly_total_charts_sheet, 2)
        closed_weekly_charts_worksheet.sheet_properties.tabColor = "1072BA"
        create_closed_weekly_totals_pivot_tables(workBook, closed_weekly_pivots_worksheet)
    else:
        closed_weekly_pivots_worksheet = workBook.get_sheet_by_name(closed_weekly_total_pivots_sheet)
        closed_weekly_charts_worksheet = workBook.get_sheet_by_name(closed_weekly_total_charts_sheet)
        update_weekly_total_pivot_tables(workBook, closed_weekly_pivots_worksheet)

    chart1 = BarChart()
    chart1.height = 12
    chart1.width = 30
    chart1.style = 10
    chart1.title = "Weekly Total - All Tickets"
    chart1.y_axis.title = 'Total'
    chart1.x_axis.title = 'Run Date'
    data = Reference(closed_weekly_pivots_worksheet, min_col=2, min_row=1,
                     max_row=closed_weekly_pivots_worksheet.max_row, max_col=closed_weekly_pivots_worksheet.max_column)
    cats = Reference(closed_weekly_pivots_worksheet, min_col=1, min_row=2,
                     max_row=closed_weekly_pivots_worksheet.max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    # chart1.series[0].trendline = Trendline()
    # chart1.series[0].trendline.trendlineType = 'linear'
    # chart1.dataLabels = DataLabelList()
    # chart1.dataLabels.showVal = True

    closed_weekly_charts_worksheet.add_chart(chart1, "A1")

    chart2 = BarChart()
    chart2.height = 12
    chart2.width = 30
    chart2.style = 10
    chart2.title = "Weekly Total - All Tickets"
    chart2.y_axis.title = 'Total'
    chart2.x_axis.title = 'Run Date'
    data = Reference(closed_weekly_pivots_worksheet, min_col=3, min_row=1,
                     max_row=closed_weekly_pivots_worksheet.max_row, max_col=3)
    cats = Reference(closed_weekly_pivots_worksheet, min_col=1, min_row=2,
                     max_row=closed_weekly_pivots_worksheet.max_row)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.shape = 4
    # chart1.series[0].trendline = Trendline()
    # chart1.series[0].trendline.trendlineType = 'linear'
    # chart1.dataLabels = DataLabelList()
    # chart1.dataLabels.showVal = True

    closed_weekly_charts_worksheet.add_chart(chart2, "A20")


def extract_data_from_old_file_and_insert_into_new_file():
    for filename in os.listdir(CURRENT_DIRECTORY):
        match = re.match('(From.*?.xlsm)', filename, re.I)
        if match:
            old_workbook_file_name = match.group(0)
    old_workbook_file_name = os.path.join(CURRENT_DIRECTORY, old_workbook_file_name)
    if os.path.exists(old_workbook_file_name):
        ertProjects = old_work_book_project_name_mapper.values()
        print "Loading the old workbook {}".format(old_workbook_file_name)
        oldWorkBook = load_workbook(old_workbook_file_name, data_only=True)
        rollUpSheet = oldWorkBook['Rollup']
        oldData = collections.OrderedDict()
        RollupData = collections.OrderedDict()
        print "Extracting Data from {} workbook and loading into {} workbook".format(old_workbook_file_name, excelFileName)
        for row in rollUpSheet.iter_rows():
            if row[0].row == 1:
                continue
            oldData[row[0].value + '##' + str(row[1].value)] = "##".join(
                [str(row[2].value), str(row[5].value), str(row[8].value)])
            data = []
            for i in range(2, 17):
                data.append(row[i].value)
            data = [str(x) for x in data]
            RollupData[row[0].value + '##' + str(row[1].value)] = "##".join(data)
        dates = set()
        keys = oldData.keys()
        for key in keys:
            project, date = key.split("##")
            if is_date(date):
                dates.add(date)
        dates = sorted(dates, key=lambda x: datetime.datetime.strptime(x, '%Y-%m-%d %H:%M:%S'))
        latestWorkBook = load_workbook(excelFileName)
        # populate the project specific sheets
        for project in ertProjects:
            project1 = project
            if project1 == 'Expert':
                project1 = 'EXPRT'
            elif project1 == 'ePRO':
                project1 = 'EPR'
            ws = latestWorkBook[project1.upper()]
            for date in dates:
                key = project + '##' + date
                if key in oldData.keys():
                    (New, InProgress, Closed) = oldData[key].split('##')
                    project_sheet_max_row = ws.max_row
                    date1 = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S').strftime("%m/%d/%Y")
                    weekOfYear = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S').strftime("%W-%Y")
                    if project_sheet_max_row == 1:
                        diff1 = diff2 = diff3 = 0
                    else:
                        diff1 = "=C{0}-C{1}".format(project_sheet_max_row + 1, project_sheet_max_row)
                        diff2 = "=E{0}-E{1}".format(project_sheet_max_row + 1, project_sheet_max_row)
                        diff3 = "=G{0}-G{1}".format(project_sheet_max_row + 1, project_sheet_max_row)
                    row = [weekOfYear, date1, int(New), diff1, int(InProgress), diff2, int(Closed), diff3]
                    ws.append(row)
        ##populate the Rollup
        latestRollup = latestWorkBook['Rollup']
        for date in dates:
            for project_code, project in old_work_book_project_name_mapper.iteritems():
                key = project + '##' + date
                project1 = latest_project_name_mapper[project_code]
                if key in RollupData.keys():
                    date1 = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S').strftime("%m/%d/%Y")
                    values = RollupData[key].split('##')
                    int_values = []
                    for value in values:
                        if value == 'none':
                            value = 0
                        try:
                            value = int(value)
                        except Exception:
                            value = str(value)
                        int_values.append(value)
                    row = [project1, date1] + int_values
                    latestRollup.append(row)

        latestWorkBook.save(filename=excelFileName)
        print "Data Extraction Completed"
    else:
        print "{} File Not exists".format(old_workbook_file_name)


def create_weekly_total_pivot_tables(workBook, pivots_worksheet):
    metrics = collections.OrderedDict()
    for project in ertProjects:
        ws = workBook[project]
        for row in ws.iter_rows():
            rowList = []
            for cell in row:
                if cell.row == 1:
                    continue
                rowList.append(cell.value)
            if rowList:
                metrics[project + '#' + rowList[1]] = '#'.join(str(v) for v in rowList[2:])
    keys = metrics.keys()
    rundatesSet = set()
    rundates = list()
    for key in keys:
        project, rundate = key.split('#')
        if rundate not in rundatesSet:
            rundatesSet.add(rundate)
            rundates.append(rundate)
    project_rows = []
    for rundate in rundates:
        total = 0
        for project in ertProjects:
            (New, diff1, InProgess, diff2, closed, diff3) = metrics[project + '#' + rundate].split('#')
            projectTotal = int(New) + int(InProgess) + int(closed)
            total = total + projectTotal
        project_rows.append((rundate, total))
    change_in_growth = []
    for i in range(1, len(project_rows)):
        change_in_growth.append((project_rows[i][0], project_rows[i][1] - project_rows[i - 1][1]))
    change_in_growth = [('Date', 'Weekly Growth in Tickets')] + change_in_growth
    for row in change_in_growth:
        pivots_worksheet.append(row)
    col = 5
    row = 1
    project_rows = [('Date', 'Sum of All Tickets')] + project_rows
    for data_row in project_rows:
        pivots_worksheet.cell(row=row, column=col).value = data_row[0]
        pivots_worksheet.cell(row=row, column=col + 1).value = data_row[1]
        row = row + 1


def get_maximum_row(work_sheet, column_number):
    max_row = 0
    for row_num in range(1, work_sheet.max_row + 2):
        if not work_sheet.cell(row=row_num, column = column_number).value:
            max_row = row_num - 1
            break
    return  max_row


def update_weekly_total_pivot_tables(workBook, pivots_worksheet):
    metrics = collections.OrderedDict()
    for project in ertProjects:
        ws = workBook[project]
        row = ws.max_row
        latest_row = []
        for col in range(ws.min_column, ws.max_column + 1):
            latest_row.append((ws.cell(row=row, column=col).value))
        metrics[project + '#' + latest_row[1]] = '#'.join(str(v) for v in latest_row[2:])
    project, rundate = metrics.keys()[0].split('#')
    total_tickets_count = 0
    for project in ertProjects:
        key = project + '#' + rundate
        (New, diff1, InProgess, diff2, closed, diff3) = metrics[key].split('#')
        projectTotal = int(New) + int(InProgess) + int(closed)
        total_tickets_count = (total_tickets_count + projectTotal)
    newRow = (rundate, total_tickets_count)


    weekly_growth_max_row = get_maximum_row(pivots_worksheet, 1)
    weekly_total_max_row = get_maximum_row(pivots_worksheet, 5)
    weekly_growth_updated = False
    for row in range(2, weekly_growth_max_row + 1):
        run_date = pivots_worksheet.cell(row=row, column=1).value
        weekly_growth = pivots_worksheet.cell(row=row, column=2).value
        print run_date, weekly_growth
        print type(run_date), run_date
        if type(run_date) == datetime.datetime:
            run_date = run_date.strftime("%m/%d/%Y")
        else:
            run_date = datetime.datetime.strptime(run_date, "%m/%d/%Y").strftime("%m/%d/%Y")
        if run_date == currentDate:
            print "updating the weekly growth pivot metrics for {}".format(currentDate)
            previous_week_total = pivots_worksheet.cell(row=weekly_total_max_row-1, column=6).value
            change_in_growth_for_current_week = total_tickets_count - previous_week_total
            pivots_worksheet.cell(row=row, column=1, value=run_date)
            pivots_worksheet.cell(row=row, column=2, value=change_in_growth_for_current_week)
            weekly_growth_updated = True
    if not weekly_growth_updated:
        print "Adding the weekly growth pivot metrics for {}".format(currentDate)
        previous_week_total = pivots_worksheet.cell(row=weekly_total_max_row, column=6).value
        change_in_growth_for_current_week = total_tickets_count - previous_week_total
        pivots_worksheet.cell(row=weekly_growth_max_row + 1, column=1, value=currentDate)
        pivots_worksheet.cell(row= weekly_growth_max_row + 1, column=2, value=change_in_growth_for_current_week)
    weekly_total_updated = False
    for row in range(2, weekly_total_max_row + 1):
        run_date = pivots_worksheet.cell(row=row, column=5).value
        if type(run_date) == datetime.datetime:
            run_date = run_date.strftime("%m/%d/%Y")
        else:
            run_date = datetime.datetime.strptime(run_date, "%m/%d/%Y").strftime("%m/%d/%Y")
        if run_date == currentDate:
            print "updating the weekly total pivot metrics for {}".format(currentDate)
            pivots_worksheet.cell(row=row, column=5, value=run_date)
            pivots_worksheet.cell(row=row, column=6, value=total_tickets_count)
            weekly_total_updated = True
    if not weekly_total_updated:
        print "Adding the weekly total pivot metrics for {}".format(currentDate)
        pivots_worksheet.cell(row=weekly_total_max_row + 1, column=5, value=currentDate)
        pivots_worksheet.cell(row= weekly_total_max_row + 1, column=6, value=total_tickets_count)


def create_or_update_weekly_total_charts(excelFileName, currentDate):
    (weekly_total_pivots_sheet, weekly_total_charts_sheet) = ('Pivot-WeeklyTotals', 'Charts-WeeklyTotals')
    workBook = load_workbook(excelFileName)
    sheets = workBook.get_sheet_names()
    if not (weekly_total_pivots_sheet in sheets and weekly_total_charts_sheet in sheets):
        print "Creating Sheet {}".format(weekly_total_pivots_sheet)
        pivots_worksheet = workBook.create_sheet(weekly_total_pivots_sheet, 0)
        pivots_worksheet.sheet_properties.tabColor = "1072BA"
        print "Creating Sheet {}".format(weekly_total_charts_sheet)
        charts_worksheet = workBook.create_sheet(weekly_total_charts_sheet, 0)
        charts_worksheet.sheet_properties.tabColor = "1072BA"
        create_weekly_total_pivot_tables(workBook, pivots_worksheet)
    else:
        pivots_worksheet = workBook.get_sheet_by_name(weekly_total_pivots_sheet)
        charts_worksheet = workBook.get_sheet_by_name(weekly_total_charts_sheet)
        update_weekly_total_pivot_tables(workBook, pivots_worksheet)

    workBook.save(filename=excelFileName)
    growth_change_date_column_number = 1
    growth_change_value_column_number = 2
    weekly_total_date_column_number = 5
    weekly_total_value_column_number = 6
    weekly_growth_max_row = get_maximum_row(pivots_worksheet, 1)
    weekly_total_max_row = get_maximum_row(pivots_worksheet, 5)

    chart1 = BarChart()
    chart1.height = 12
    chart1.width = 30
    chart1.style = 10
    chart1.title = "Weekly Total - All Tickets"
    chart1.y_axis.title = 'Total'
    chart1.x_axis.title = 'Run Date'
    data = Reference(pivots_worksheet, min_col=weekly_total_value_column_number, min_row=1, max_row=weekly_total_max_row, max_col=weekly_total_value_column_number)
    cats = Reference(pivots_worksheet, min_col=weekly_total_date_column_number, min_row=2, max_row=weekly_total_max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    chart1.series[0].trendline = Trendline()
    chart1.series[0].trendline.trendlineType = 'linear'
    chart1.dataLabels = DataLabelList()
    chart1.dataLabels.showVal = True
    charts_worksheet.add_chart(chart1, "A2")

    c1 = LineChart()
    c1.height = 12
    c1.width = 30
    c1.title = "Weekly Growth"
    c1.style = 12
    c1.y_axis.title = 'Growth'
    c1.x_axis.title = 'Run Date'
    data = Reference(
        pivots_worksheet, min_col=growth_change_value_column_number, min_row=1,
        max_col=growth_change_value_column_number, max_row=weekly_growth_max_row
    )
    cats = Reference(
        pivots_worksheet, min_col=growth_change_date_column_number, min_row=2, max_row=weekly_growth_max_row
    )
    c1.add_data(data, titles_from_data=True)
    c1.set_categories(cats)
    # Style the lines
    s1 = c1.series[0]
    s1.marker.symbol = "circle"
    s1.marker.graphicalProperties.solidFill = "360AD2"  # Marker filling
    s1.marker.graphicalProperties.line.solidFill = "360AD2"
    s1.graphicalProperties.line.solidFill = "360AD2"
    s1.graphicalProperties.line.width = 28568  # width in EMUs
    s1.trendline = Trendline()
    s1.trendline.trendlineType = 'linear'
    # s1.smooth = True # Make the line smooth

    charts_worksheet.add_chart(c1, "A30")
    #workBook.save(filename=excelFileName)


if __name__ == '__main__':
    CURRENT_DIRECTORY = os.path.dirname(os.path.realpath(__file__))
    config_file = os.path.join(CURRENT_DIRECTORY, 'config', 'jiraMetrics.ini')
    if os.path.exists(config_file):
        config = SafeConfigParser()
        config.read(config_file)
    else:
        print config_file + " not found"
        time.sleep(5)
        sys.exit(0)

    jira_api_obj = JiraAPIHandler(config)

    output_dir = os.path.join(CURRENT_DIRECTORY, 'output')
    excelFileName = os.path.join(output_dir, config.get('OUTPUT', 'output_file_name'))

    old_work_book_project_name_mapper = collections.OrderedDict()
    latest_project_name_mapper = collections.OrderedDict()
    project_code_vs_names = config.get('BUG_TRACKER', 'old_workbook_project_code_vs_name_map').split(',')
    for item in project_code_vs_names:
        (project_code, project_name) = item.split('=>')
        old_work_book_project_name_mapper[project_code.strip()] = project_name.strip()

    latest_project_code_vs_names = config.get('BUG_TRACKER', 'project_code_vs_name_map').split(',')
    for item in latest_project_code_vs_names:
        (project_code, project_name) = item.split('=>')
        latest_project_name_mapper[project_code.strip()] = project_name.strip()

    jsonOutputDir = os.path.join(CURRENT_DIRECTORY, 'json')
    excelOutputDir = os.path.join(CURRENT_DIRECTORY, 'excel')

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    if not os.path.exists(jsonOutputDir):
        os.makedirs(jsonOutputDir)
    if not os.path.exists(excelOutputDir):
        os.makedirs(excelOutputDir)

    if os.path.exists(excelFileName):
        try:
            os.rename(excelFileName, excelFileName)
        except OSError as e:
            print excelFileName + ' already in use. Please close it'
            time.sleep(5)
            sys.exit(0)

    ertProjects = [project.strip() for project in config.get('BUG_TRACKER', 'projects').split(",")]

    status_arr = []
    jql_items = config.items('JQL')
    for status, query in jql_items:
        status_arr.append(status)
    status_arr = [string.capwords(status) for status in status_arr]
    days_to_subtract = config.get('BUG_TRACKER', 'day_difference')
    try:
        days_to_subtract = int(days_to_subtract)
    except ValueError:
        days_to_subtract = 0

    currentDate = datetime.date.today() - datetime.timedelta(days=days_to_subtract)
    currentWeek = currentDate.strftime("%W-%Y")
    currentDate_YYYY_MM_DD = currentDate.strftime("%Y-%m-%d")
    currentDate = currentDate.strftime("%m/%d/%Y")

    # Create Empty worksheets if the file not exists
    if not os.path.exists(excelFileName):
        print "Creating empty workbook as {}".format(excelFileName)
        statusList = ['Current Week', 'Last Week', 'Difference']
        finalStatusList = ['Current Week', 'Last Week', 'Growth']
        row = '2'
        workBook = Workbook()
        ws = workBook.active
        # Rollup Section setup ####
        ws.title = "Rollup"
        # setting Project ###
        ws.merge_cells('A1:A2')
        ws['A1'] = "Project"
        ws.merge_cells('B1:B2')
        ws['B1'] = "Run Date"
        # setting New section ###
        ws.merge_cells('C1:E1')
        ws['C1'].value = "New"

        for i, j in zip(range(3), range(ord('C'), ord('E') + 1)):
            ws[chr(j) + row] = statusList[i]
        # setting In Progress section ###
        ws.merge_cells('F1:H1')
        ws['F1'].value = "In Progress"
        for i, j in zip(range(3), range(ord('F'), ord('H') + 1)):
            ws[chr(j) + row] = statusList[i]
        # setting Closed section #
        ws.merge_cells('I1:K1')
        ws['I1'].value = "Closed"
        for i, j in zip(range(3), range(ord('I'), ord('K') + 1)):
            ws[chr(j) + row] = statusList[i]
        # setting New & In progress section #
        ws.merge_cells('L1:N1')
        ws['L1'].value = "New & In Progress"
        for i, j in zip(range(3), range(ord('L'), ord('N') + 1)):
            ws[chr(j) + row] = statusList[i]
        # setting Total section #
        ws.merge_cells('O1:Q1')
        ws['O1'].value = "Total"
        for i in range(ord('O'), ord('Q') + 1):
            ws[chr(i) + row] = finalStatusList.pop(0)
        ws.freeze_panes = 'A3'
        sheet_cols = status_arr
        for index in (1, 3, 5):
            sheet_cols.insert(index, 'diff')
        row = ['Week#', 'Run Date'] + sheet_cols

        for project in ertProjects:
            workSheet = workBook.create_sheet(project)
            workSheet.append(row)
            workSheet.freeze_panes = 'A2'
        print "Workbook created"
        print "Applying Styles to workbook"
        header_font = Font(name='Calibri', size=12, bold=True)
        side = Side(border_style='thin', color="FF000000")
        color_fill = PatternFill("solid", fgColor="87CEEB")
        wrap_alignment = Alignment(wrap_text=True, vertical="top", horizontal='center')
        for sheet in workBook.get_sheet_names():
            ws = workBook[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom
                    )
                    border.left = border.right = border.top = border.bottom = side
                    cell.alignment = wrap_alignment
                    cell.border = border
                    cell.font = header_font
                    cell.fill = color_fill
        workBook.save(filename=excelFileName)
        print "Styles applied"
        extract_data_from_old_file_and_insert_into_new_file()

    print "Getting the latest metrics from jira for {}".format(currentDate)
    workBook = load_workbook(excelFileName)
    # populating data for project Sheets #
    projectWorkSheet = workBook[ertProjects[0]]
    lastRunWeek = projectWorkSheet.cell(row=projectWorkSheet.max_row, column=projectWorkSheet.min_column).value
    lastRunDate = projectWorkSheet.cell(row=projectWorkSheet.max_row, column=projectWorkSheet.min_column + 1).value
    lastRunDate = datetime.datetime.strptime(lastRunDate, "%m/%d/%Y").strftime("%m/%d/%Y")
    script_executed_for_current_week = (projectWorkSheet.max_row > 1 and (lastRunDate == currentDate))
    '''
    script_executed_for_current_week = (
        projectWorkSheet.max_row > 1 and (lastRunWeek == currentWeek) or (lastRunDate==currentDate)
    )
    '''
    rollUpSheet = workBook['Rollup']
    rollUpSheet_max_row = rollUpSheet.max_row
    if script_executed_for_current_week:
        rollUpSheet_max_row = (rollUpSheet_max_row - len(ertProjects))

    rollupSheetRows = []
    rollupIndex = 1
    lastWeekResults = dict()
    currentWeekResults = dict()
    for project in ertProjects:
        print "working on Metrics for {} project ".format(project)
        workSheet = workBook[project]
        lastWeekResults = dict()
        currentWeekResults = dict()
        row = []
        projectRowIndex = 1
        if config.has_section(project + '_JQL'):
            jqlQueries = collections.OrderedDict(config.items(project + '_JQL'))
        else:
            jqlQueries = collections.OrderedDict(jql_items)
        for status, query in jqlQueries.iteritems():
            status = string.capwords(status)
            query = query.replace('__PROJECTNAME__', project)
            query = query.replace('__CURRENTDATE__', currentDate_YYYY_MM_DD)
            response = jira_api_obj.get_response_from_jira(query)
            queryCount = response['total']
            # time.sleep(10)
            row.append(queryCount)
            print status, queryCount
            currentWeekResults[project + '-' + status] = queryCount
            project_sheet_max_row = workSheet.max_row
            if script_executed_for_current_week:
                project_sheet_max_row = (project_sheet_max_row - 1)

            if project_sheet_max_row == 1:
                lastRunValue = 0
                diff = 0
            elif status == 'New':
                lastRunValue = workSheet['C' + str(project_sheet_max_row)].value
                diff = "=C{0}-C{1}".format(project_sheet_max_row + 1, project_sheet_max_row)
            elif status == 'In Progress':
                lastRunValue = workSheet['E' + str(project_sheet_max_row)].value
                diff = "=E{0}-E{1}".format(project_sheet_max_row + 1, project_sheet_max_row)
            elif status == 'Closed':
                lastRunValue = workSheet['G' + str(project_sheet_max_row)].value
                diff = "=G{0}-G{1}".format(project_sheet_max_row + 1, project_sheet_max_row)
            row.append(diff)
            lastWeekResults[project + '-' + status] = lastRunValue
        row = [currentWeek, currentDate] + row
        if script_executed_for_current_week:
            index = 0
            for col in range(workSheet.min_column, workSheet.max_column + 1):
                workSheet.cell(row=project_sheet_max_row + 1, column=col, value=row[index])
                index += 1
        else:
            workSheet.append(row)
        if rollUpSheet_max_row == 2:
            currentWeekTotal = currentWeekResults[project + '-New'] +\
                currentWeekResults[project + '-In Progress'] + currentWeekResults[project + '-Closed']
            rollupSheetRows.append([latest_project_name_mapper[project], currentDate,
                                    currentWeekResults[project + '-New'],
                                    lastWeekResults[project + '-New'],
                                    0,
                                    currentWeekResults[project + '-In Progress'],
                                    lastWeekResults[project + '-In Progress'],
                                    0,
                                    currentWeekResults[project + '-Closed'],
                                    lastWeekResults[project + '-Closed'],
                                    0,
                                    "=C{0}+F{0}".format(rollUpSheet_max_row + rollupIndex),
                                    0, 0,
                                    "=I{0}+L{0}".format(rollUpSheet_max_row + rollupIndex),
                                    0, 0
                                    ])
        else:
            rollupSheetRows.append([latest_project_name_mapper[project], currentDate,
                                    currentWeekResults[project + '-New'],
                                    lastWeekResults[project + '-New'],
                                    "=C{0}-D{0}".format(rollUpSheet_max_row + rollupIndex),
                                    currentWeekResults[project + '-In Progress'],
                                    lastWeekResults[project + '-In Progress'],
                                    "=F{0}-G{0}".format(rollUpSheet_max_row + rollupIndex),
                                    currentWeekResults[project + '-Closed'],
                                    lastWeekResults[project + '-Closed'],
                                    "=I{0}-J{0}".format(rollUpSheet_max_row + rollupIndex),
                                    "=C{0}+F{0}".format(rollUpSheet_max_row + rollupIndex),
                                    "=D{0}+G{0}".format(rollUpSheet_max_row + rollupIndex),
                                    "=L{0}-M{0}".format(rollUpSheet_max_row + rollupIndex),
                                    "=I{0}+L{0}".format(rollUpSheet_max_row + rollupIndex),
                                    "=J{0}+M{0}".format(rollUpSheet_max_row + rollupIndex),
                                    "=O{0}-P{0}".format(rollUpSheet_max_row + rollupIndex),
                                    ])
        rollupIndex += 1
    # populate data for Rollup Sheet ###
    if script_executed_for_current_week:
        index = 1
        for row in rollupSheetRows:
            for col in range(rollUpSheet.min_column, rollUpSheet.max_column + 1):
                rollUpSheet.cell(row=rollUpSheet_max_row + index, column=col, value=rollupSheetRows[index-1][col-1])
            index += 1
    else:
        for row in rollupSheetRows:
            rollUpSheet.append(row)
    workBook.save(filename=excelFileName)
    old_weekly_totals_sheet = 'WeeklyTotals'
    if old_weekly_totals_sheet in workBook.get_sheet_names():
        std = workBook.get_sheet_by_name(old_weekly_totals_sheet)
        workBook.remove_sheet(std)
    workBook.save(filename=excelFileName)
    create_or_update_weekly_total_charts(excelFileName, currentDate)
    create_or_update_closed_weekly_total_charts(excelFileName, currentDate)
    workBook.save(filename=excelFileName)
    print "Task Completed"
